"""Core report generation logic - data aggregation and export to PDF, Excel, CSV, HTML."""
from __future__ import annotations

import csv
import io
import logging
import re as _re
from datetime import datetime, timezone

import json
import markdown as _md

from jinja2 import Environment, FileSystemLoader

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

# Regex matching characters illegal in Excel/openpyxl cells (ASCII control chars except tab/newline/CR)
_ILLEGAL_XLSX_RE = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

from backend.ai.llm_manager import llm_manager
from backend.core.false_positive_detector import enrich_findings_with_fp
from backend.core.text_utils import normalize_unicode, smart_truncate
from backend.models.benchmark import Benchmark
from backend.models.client import Client
from backend.models.discovery_cache import DiscoveryCache
from backend.models.finding import Finding
from backend.models.mission import Mission
from backend.models.rule import Rule
from backend.models.scan import Scan
from backend.models.target import Target

logger = logging.getLogger("auditforge.reports")

TEMPLATES_DIR = str(__import__("pathlib").Path(__file__).resolve().parent.parent / "templates")

# Project logo (transparent 128×128 PNG) embedded as base64 for PDF cover
_PROJECT_LOGO_PATH = __import__("pathlib").Path(__file__).resolve().parent / "_project_logo_b64.txt"
_PROJECT_LOGO_B64 = _PROJECT_LOGO_PATH.read_text().strip() if _PROJECT_LOGO_PATH.exists() else ""


def _md_to_html(text: str) -> str:
    """Convert Markdown text (from AI summary) to safe HTML for embedding in reports."""
    if not text:
        return ""
    return _md.markdown(text, extensions=["tables", "nl2br"])

# Excel color fills
FILL_CRITICAL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_HIGH = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
FILL_MEDIUM = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
FILL_LOW = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
FILL_INFORMATIONAL = PatternFill(start_color="9CA3AF", end_color="9CA3AF", fill_type="solid")
FILL_PASS = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

SEVERITY_FILLS = {
    "critical": FILL_CRITICAL,
    "high": FILL_HIGH,
    "medium": FILL_MEDIUM,
    "low": FILL_LOW,
    "informational": FILL_INFORMATIONAL,
}


# ---------------------------------------------------------------------------
# Service-to-port keyword mapping for finding ↔ discovery linkage
# ---------------------------------------------------------------------------

# Maps keywords (found in rule title/description) → sets of port numbers.
# When a finding's rule text matches a keyword AND the target has that port
# open in its discovery data, an informational note is attached.
_SERVICE_PORT_MAP: dict[str, set[int]] = {
    "ssh": {22},
    "rdp": {3389},
    "remote desktop": {3389},
    "winrm": {5985, 5986},
    "smb": {139, 445},
    "server message block": {139, 445},
    "netbios": {135, 137, 139},
    "telnet": {23},
    "ftp": {21},
    "http": {80, 8080, 8000, 8443, 8888, 9090},
    "https": {443, 8443},
    "dns": {53},
    "ldap": {389, 636},
    "kerberos": {88},
    "snmp": {161},
    "smtp": {25},
    "imap": {143},
    "pop3": {110},
    "vnc": {5900},
    "mssql": {1433},
    "sql server": {1433},
    "mysql": {3306},
    "postgresql": {5432},
    "oracle": {1521},
    "redis": {6379},
    "mongodb": {27017},
    "nfs": {2049},
    "sip": {5060},
    "syslog": {514},
    "printing": {631, 9100},
    "printer": {631, 9100},
}


# ---------------------------------------------------------------------------
# Service Risk Knowledge Base — static, deterministic, no AI required.
# Maps port number → human-readable service metadata + generic risk text.
# ---------------------------------------------------------------------------

_SERVICE_RISK_KB: dict[int, dict] = {
    21: {
        "name": "FTP (File Transfer Protocol)",
        "purpose": "Unencrypted file transfer between systems",
        "risk_when_rule_fails": "FTP transmits credentials and data in cleartext. If this control is misconfigured and FTP is exposed, attackers can intercept credentials via network sniffing or brute-force FTP logins to access sensitive files.",
        "attack_examples": "Credential sniffing, anonymous FTP access, brute-force attacks",
        "business_impact": "Unauthorized file access, data leakage, credential theft",
    },
    22: {
        "name": "SSH (Secure Shell)",
        "purpose": "Encrypted remote command-line access to servers and network devices",
        "risk_when_rule_fails": "If SSH is exposed and this security control is misconfigured, attackers could exploit weak authentication, outdated ciphers, or missing access controls to gain remote shell access with full command execution capability.",
        "attack_examples": "Brute-force login, credential stuffing, CVE-2024-6387 (regreSSHion)",
        "business_impact": "Full system compromise, unauthorized data access, lateral movement into the internal network",
    },
    23: {
        "name": "Telnet",
        "purpose": "Unencrypted remote terminal access (legacy protocol)",
        "risk_when_rule_fails": "Telnet sends all data including credentials in cleartext. Combined with a misconfigured security control, this allows trivial credential interception and unauthorized system access.",
        "attack_examples": "Cleartext credential capture, man-in-the-middle attacks, brute-force login",
        "business_impact": "Complete credential exposure, unauthorized system control, compliance violations",
    },
    25: {
        "name": "SMTP (Simple Mail Transfer Protocol)",
        "purpose": "Email transmission between mail servers",
        "risk_when_rule_fails": "Exposed SMTP with misconfigured controls enables open relay abuse, phishing email injection, and email header manipulation for social engineering attacks.",
        "attack_examples": "Open relay abuse, phishing campaigns, email spoofing",
        "business_impact": "Reputation damage from spam relaying, phishing attacks targeting employees and partners",
    },
    53: {
        "name": "DNS (Domain Name System)",
        "purpose": "Name resolution translating domain names to IP addresses",
        "risk_when_rule_fails": "Exposed DNS with security misconfigurations enables DNS cache poisoning, zone transfer leakage, and DNS amplification attacks that can redirect traffic or leak internal network topology.",
        "attack_examples": "DNS cache poisoning, zone transfer enumeration, DNS amplification DDoS",
        "business_impact": "Traffic interception, internal network mapping exposure, service disruption",
    },
    80: {
        "name": "HTTP (Web Server)",
        "purpose": "Unencrypted web service delivery",
        "risk_when_rule_fails": "Unencrypted HTTP combined with failed security controls exposes web applications to session hijacking, credential theft, and content injection attacks.",
        "attack_examples": "Session hijacking, credential interception, cross-site scripting",
        "business_impact": "Data theft via web applications, customer-facing service compromise",
    },
    88: {
        "name": "Kerberos",
        "purpose": "Active Directory authentication protocol",
        "risk_when_rule_fails": "Kerberos exposure with misconfigured authentication controls enables Kerberoasting, AS-REP roasting, and ticket-based attacks that can compromise domain credentials without brute-forcing.",
        "attack_examples": "Kerberoasting, AS-REP roasting, Golden Ticket, Silver Ticket attacks",
        "business_impact": "Active Directory domain compromise, privilege escalation to Domain Admin, full organizational takeover",
    },
    110: {
        "name": "POP3 (Post Office Protocol)",
        "purpose": "Email retrieval from mail servers (often unencrypted)",
        "risk_when_rule_fails": "POP3 without TLS transmits email credentials in cleartext. Combined with security misconfigurations, this enables credential capture and unauthorized mailbox access.",
        "attack_examples": "Credential sniffing, brute-force attacks, mailbox enumeration",
        "business_impact": "Email account compromise, sensitive communication exposure",
    },
    135: {
        "name": "MS-RPC (Microsoft Remote Procedure Call)",
        "purpose": "Windows inter-process communication and remote service management",
        "risk_when_rule_fails": "RPC exposure enables remote enumeration of system services, user accounts, and shares. Combined with misconfigured controls, it provides attack reconnaissance and exploitation vectors.",
        "attack_examples": "RPC endpoint enumeration, DCOM exploitation, remote service abuse",
        "business_impact": "Internal network reconnaissance, remote code execution on Windows systems",
    },
    139: {
        "name": "NetBIOS Session Service",
        "purpose": "Legacy Windows network file and printer sharing protocol",
        "risk_when_rule_fails": "NetBIOS exposure combined with misconfigured SMB/sharing policies enables null session enumeration, credential capture via NTLM relay, and exploitation of legacy protocol weaknesses.",
        "attack_examples": "Null session enumeration, NTLM relay attacks, Responder/LLMNR poisoning",
        "business_impact": "Internal network reconnaissance, credential theft, lateral movement leading to domain compromise",
    },
    143: {
        "name": "IMAP (Internet Message Access Protocol)",
        "purpose": "Email access and management on mail servers",
        "risk_when_rule_fails": "IMAP without TLS exposes email credentials during authentication. A misconfigured control increases the risk of unauthorized mailbox access and email data exfiltration.",
        "attack_examples": "Credential interception, brute-force attacks, email data exfiltration",
        "business_impact": "Email account compromise, sensitive business communication exposure",
    },
    161: {
        "name": "SNMP (Simple Network Management Protocol)",
        "purpose": "Network device monitoring and management",
        "risk_when_rule_fails": "SNMP with default or weak community strings allows attackers to read device configurations, extract credentials, and modify network device settings remotely.",
        "attack_examples": "Community string brute-force, configuration extraction, SNMP-based reconnaissance",
        "business_impact": "Network infrastructure compromise, configuration exposure, device takeover",
    },
    389: {
        "name": "LDAP (Lightweight Directory Access Protocol)",
        "purpose": "Directory services for user authentication and organizational data",
        "risk_when_rule_fails": "Unencrypted LDAP exposure with misconfigured controls allows credential capture during bind operations, directory enumeration of all user accounts, and LDAP injection attacks.",
        "attack_examples": "LDAP bind credential capture, directory enumeration, LDAP injection",
        "business_impact": "Mass credential exposure, complete Active Directory user enumeration, authentication bypass",
    },
    443: {
        "name": "HTTPS (Secure Web Server)",
        "purpose": "Encrypted web service delivery",
        "risk_when_rule_fails": "Even over TLS, web services with failed security controls may be vulnerable to certificate validation bypasses, weak cipher exploitation, or application-layer attacks.",
        "attack_examples": "TLS downgrade attacks, certificate validation bypass, web application exploitation",
        "business_impact": "Customer-facing data exposure, secure communication interception",
    },
    445: {
        "name": "SMB (Server Message Block)",
        "purpose": "Windows file sharing, printer sharing, and inter-process communication",
        "risk_when_rule_fails": "SMB exposure with a failed security control is one of the most dangerous combinations in Windows environments. Historical exploits like EternalBlue (MS17-010) and ongoing NTLM relay attacks directly target this port.",
        "attack_examples": "EternalBlue/WannaCry, PetNotPetya, NTLM relay with ntlmrelayx, CobaltStrike SMB beacon",
        "business_impact": "Organization-wide ransomware spread, mass data exfiltration from file shares, Active Directory domain takeover via relay attacks",
    },
    514: {
        "name": "Syslog",
        "purpose": "Centralized logging and event collection",
        "risk_when_rule_fails": "Exposed syslog services can be flooded to hide malicious activity or injected with false log entries to mislead incident response and forensic investigations.",
        "attack_examples": "Log injection, log flooding, evidence tampering",
        "business_impact": "Loss of audit trail integrity, impaired incident detection and response",
    },
    636: {
        "name": "LDAPS (LDAP over SSL)",
        "purpose": "Encrypted directory services for user authentication",
        "risk_when_rule_fails": "Even with TLS, misconfigurations in directory service controls can enable authentication bypass, certificate trust issues, or enumeration of directory objects.",
        "attack_examples": "Certificate validation bypass, directory enumeration over encrypted channel",
        "business_impact": "Authenticated access to directory services, user account enumeration",
    },
    1433: {
        "name": "MS-SQL (Microsoft SQL Server)",
        "purpose": "Microsoft relational database service",
        "risk_when_rule_fails": "Exposed SQL Server with security misconfigurations enables direct database access, SQL injection against stored procedures, and potential remote code execution via xp_cmdshell.",
        "attack_examples": "SQL brute-force, xp_cmdshell RCE, SQL injection, credential theft from databases",
        "business_impact": "Direct database compromise, mass data exfiltration, regulatory breach (GDPR, PCI-DSS)",
    },
    1521: {
        "name": "Oracle DB",
        "purpose": "Oracle relational database service",
        "risk_when_rule_fails": "Exposed Oracle database services with security gaps enable TNS listener attacks, database credential brute-forcing, and direct access to sensitive enterprise data.",
        "attack_examples": "TNS listener exploitation, Oracle database brute-force, privilege escalation",
        "business_impact": "Enterprise database compromise, sensitive financial/HR data exposure",
    },
    2049: {
        "name": "NFS (Network File System)",
        "purpose": "Unix/Linux network file sharing",
        "risk_when_rule_fails": "NFS exposure with misconfigured export permissions allows unauthorized mounting of file systems, data exfiltration, and potentially root-level compromise via UID spoofing.",
        "attack_examples": "NFS share enumeration, unauthorized file system mounting, UID spoofing",
        "business_impact": "Unauthorized access to shared files and directories, data leakage",
    },
    3306: {
        "name": "MySQL",
        "purpose": "MySQL relational database service",
        "risk_when_rule_fails": "Exposed MySQL with misconfigured controls enables database credential brute-forcing, unauthorized data access, and potential command execution via user-defined functions.",
        "attack_examples": "MySQL brute-force, UDF exploitation, data exfiltration via SELECT INTO OUTFILE",
        "business_impact": "Database compromise, customer data theft, application backend takeover",
    },
    3389: {
        "name": "RDP (Remote Desktop Protocol)",
        "purpose": "Graphical remote access to Windows desktops and servers",
        "risk_when_rule_fails": "With RDP exposed and this security control failing, the system is vulnerable to credential brute-forcing, pass-the-hash attacks, and exploitation of RDP protocol vulnerabilities. Attackers gain full interactive desktop access.",
        "attack_examples": "BlueKeep exploit (CVE-2019-0708), brute-force with Hydra/Crowbar, RDP session hijacking",
        "business_impact": "Ransomware deployment via remote desktop, unauthorized access to sensitive desktop applications, keylogging and screen capture",
    },
    5060: {
        "name": "SIP (Session Initiation Protocol)",
        "purpose": "Voice over IP (VoIP) call signaling",
        "risk_when_rule_fails": "Exposed SIP with security misconfigurations enables toll fraud, call interception, and denial-of-service attacks against telephony infrastructure.",
        "attack_examples": "SIP registration hijacking, toll fraud, VoIP eavesdropping",
        "business_impact": "Financial loss from toll fraud, privacy violation from call interception",
    },
    5432: {
        "name": "PostgreSQL",
        "purpose": "PostgreSQL relational database service",
        "risk_when_rule_fails": "Exposed PostgreSQL with misconfigured controls allows database brute-forcing, unauthorized data access, and potential OS command execution via COPY TO/FROM PROGRAM.",
        "attack_examples": "Database brute-force, COPY PROGRAM RCE, pg_read_file data extraction",
        "business_impact": "Database compromise, sensitive data exfiltration, backend system takeover",
    },
    5900: {
        "name": "VNC (Virtual Network Computing)",
        "purpose": "Remote graphical desktop access (often weakly authenticated)",
        "risk_when_rule_fails": "VNC frequently uses weak or no authentication. Exposure combined with security misconfigurations provides direct graphical access to the target system.",
        "attack_examples": "VNC authentication bypass, weak password brute-force, session hijacking",
        "business_impact": "Full remote desktop access, visual surveillance of user activity",
    },
    5985: {
        "name": "WinRM HTTP (Windows Remote Management)",
        "purpose": "PowerShell-based remote administration of Windows systems",
        "risk_when_rule_fails": "WinRM exposure enables remote PowerShell execution. With failed security controls, attackers can execute arbitrary commands, deploy malware, and move laterally across the Windows environment.",
        "attack_examples": "Evil-WinRM, Pass-the-Hash via WinRM, PowerShell Empire, lateral movement",
        "business_impact": "Remote code execution across Windows fleet, mass malware deployment, domain-wide compromise",
    },
    5986: {
        "name": "WinRM HTTPS (Windows Remote Management over SSL)",
        "purpose": "Encrypted PowerShell-based remote administration of Windows systems",
        "risk_when_rule_fails": "Even over TLS, WinRM with misconfigured controls allows authenticated remote PowerShell execution. Stolen or relayed credentials grant full administrative access.",
        "attack_examples": "Certificate-based WinRM abuse, credential relay, encrypted C2 channel via WinRM",
        "business_impact": "Stealthy remote administration abuse, difficult to detect lateral movement",
    },
    6379: {
        "name": "Redis",
        "purpose": "In-memory data store and cache (often unauthenticated by default)",
        "risk_when_rule_fails": "Redis defaults to no authentication. Exposure allows data extraction, key manipulation, and can escalate to remote code execution via Lua scripting or module loading.",
        "attack_examples": "Unauthenticated data access, Redis RCE via Lua eval, SSH key injection via CONFIG SET",
        "business_impact": "Cache poisoning, session hijacking, server-side RCE",
    },
    8080: {
        "name": "HTTP Alternate (Web Proxy/Application)",
        "purpose": "Alternative web service or application server port",
        "risk_when_rule_fails": "Alternate HTTP ports often run management interfaces, development servers, or proxies with weaker security controls than production web servers.",
        "attack_examples": "Admin interface exploitation, default credentials on management consoles",
        "business_impact": "Administrative access to web infrastructure, application compromise",
    },
    8443: {
        "name": "HTTPS Alternate",
        "purpose": "Alternative encrypted web service or management port",
        "risk_when_rule_fails": "Often used for management consoles (vSphere, iLO, DRAC). Misconfigured controls may allow unauthorized access to hypervisor or hardware management interfaces.",
        "attack_examples": "Management console exploitation, default credential attacks",
        "business_impact": "Hypervisor/hardware takeover, infrastructure-wide compromise",
    },
    9090: {
        "name": "Web Management Console",
        "purpose": "Administrative web interfaces for servers and appliances",
        "risk_when_rule_fails": "Management consoles on port 9090 (Cockpit, Prometheus, etc.) with failed security controls may expose system administration capabilities to unauthorized users.",
        "attack_examples": "Default credential attacks, admin panel exploitation",
        "business_impact": "System administration takeover, configuration manipulation",
    },
    9100: {
        "name": "Network Printing (JetDirect)",
        "purpose": "Direct network printing protocol",
        "risk_when_rule_fails": "Exposed print services can be abused for data exfiltration via print jobs, printer memory extraction, and as pivot points into segmented networks.",
        "attack_examples": "PRET exploitation, print job interception, printer-based pivoting",
        "business_impact": "Document data leakage, network segmentation bypass",
    },
    27017: {
        "name": "MongoDB",
        "purpose": "NoSQL document database service",
        "risk_when_rule_fails": "MongoDB historically defaults to no authentication. Exposure allows direct database access, data exfiltration, and potential ransomware-style data deletion.",
        "attack_examples": "Unauthenticated database access, MongoDB ransom attacks, data exfiltration",
        "business_impact": "Mass data theft or destruction, customer data exposure, regulatory breach",
    },
}


def _safe_json_loads(raw: str, default):
    """Parse a JSON string, returning *default* on any error."""
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError, ValueError):
        return default


def _parse_os_details(raw: str) -> dict:
    """Parse target.os_details which may be a JSON dict or a plain string."""
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return parsed
    except (json.JSONDecodeError, TypeError):
        pass
    return {"os": raw}


def _build_os_display(os_guess: str, os_version: str) -> str:
    """Build a clean, non-redundant OS display string from guess + version.

    Avoids duplication like "Windows Server Windows Server 10" by detecting
    when os_version already contains the os_guess text.
    """
    g = os_guess.strip()
    v = os_version.strip()
    if not g and not v:
        return ""
    if not v:
        return g
    if not g:
        return v
    # If os_version already starts with os_guess, just use os_version
    if v.lower().startswith(g.lower()):
        return v
    # If os_guess is fully contained in os_version, just use os_version
    if g.lower() in v.lower():
        return v
    return f"{g} {v}"


def _build_device_profiles(
    targets_map: dict[int, dict],
    _targets_bulk: dict,
    db: Session,
) -> dict[int, dict]:
    """Build a device profile for every target.

    Uses DiscoveryCache when available (full port/OS data).  Falls back to
    the target's own ``os_details`` JSON so that device profiles always
    appear in technical reports even before a port scan has been run.
    """
    profiles: dict[int, dict] = {}

    for tid, tdata in targets_map.items():
        target_obj = _targets_bulk.get(tid)
        if not target_obj:
            continue

        # Look up by MAC first (most reliable), then by IP
        cache_row: DiscoveryCache | None = None
        if target_obj.mac_address:
            cache_row = (
                db.query(DiscoveryCache)
                .filter(DiscoveryCache.mac_address == target_obj.mac_address)
                .order_by(DiscoveryCache.last_seen.desc())
                .first()
            )
        if not cache_row and target_obj.ip_address:
            cache_row = (
                db.query(DiscoveryCache)
                .filter(DiscoveryCache.ip_address == target_obj.ip_address)
                .order_by(DiscoveryCache.last_seen.desc())
                .first()
            )

        if cache_row:
            open_ports = cache_row.open_ports  # parsed JSON list
            _os_g = (cache_row.os_guess or "").strip()
            _os_v = (cache_row.os_version or "").strip()
            _os_display = _build_os_display(_os_g, _os_v)
            profiles[tid] = {
                "hostname": cache_row.hostname or tdata.get("hostname", ""),
                "ip_address": cache_row.ip_address or tdata.get("ip_address", ""),
                "mac_address": cache_row.mac_address or "",
                "os_guess": _os_g,
                "os_version": _os_v,
                "os_display": _os_display,
                "vendor": cache_row.vendor or "",
                "device_model": cache_row.device_model or "",
                "firmware": cache_row.firmware or "",
                "domain": cache_row.domain or "",
                "detection_method": cache_row.detection_method or "",
                "confidence": cache_row.confidence or 0,
                "open_ports": open_ports,
                "connection_methods": cache_row.connection_methods,
                "first_seen": cache_row.first_seen.isoformat() if cache_row.first_seen else "",
                "last_seen": cache_row.last_seen.isoformat() if cache_row.last_seen else "",
                "has_data": True,
                "has_port_data": bool(open_ports),
            }
        else:
            # Fallback: build a basic profile from Target's own data
            os_info = _parse_os_details(target_obj.os_details or "")
            _os_g = os_info.get("os", "")
            _os_v = os_info.get("os_version", "")
            profiles[tid] = {
                "hostname": tdata.get("hostname", "") or target_obj.hostname or "",
                "ip_address": tdata.get("ip_address", "") or target_obj.ip_address or "",
                "mac_address": (target_obj.mac_address or ""),
                "os_guess": _os_g,
                "os_version": _os_v,
                "os_display": _build_os_display(_os_g, _os_v),
                "vendor": "",
                "device_model": "",
                "firmware": "",
                "domain": "",
                "detection_method": "target_info",
                "confidence": 0,
                "open_ports": [],
                "connection_methods": [],
                "first_seen": "",
                "last_seen": "",
                "has_data": True,
                "has_port_data": False,
            }

    return profiles


# ---------------------------------------------------------------------------
# Port entry enrichment — backfill service names from the Risk KB
# ---------------------------------------------------------------------------

# Reverse lookup: port number → canonical service name from _SERVICE_RISK_KB
_PORT_TO_SERVICE: dict[int, str] = {
    port: meta["name"].split("(")[0].strip()
    for port, meta in _SERVICE_RISK_KB.items()
}


def _enrich_port_entries(profiles: dict[int, dict]) -> None:
    """Enrich open_ports entries that are missing service/product/banner.

    Operates in-place.  For each port entry lacking a 'service' key (or
    having an empty one), fills in service name, purpose, and a concise
    info line from ``_SERVICE_RISK_KB``.
    """
    for profile in profiles.values():
        for p in profile.get("open_ports", []):
            port_num = p.get("port")
            if not port_num:
                continue
            kb = _SERVICE_RISK_KB.get(port_num)
            # Fill missing service name
            if not p.get("service"):
                if kb:
                    p["service"] = kb["name"].split("(")[0].strip()
                else:
                    p["service"] = f"PORT-{port_num}"
            # Fill missing banner/info with purpose
            if not p.get("banner_snippet") and not p.get("banner") and kb:
                p["banner"] = kb["purpose"]


def _link_findings_to_services(
    findings_rows: list[dict],
    device_profiles: dict[int, dict],
    scans_map: dict[int, "Scan"],
) -> None:
    """Annotate each finding with enriched port-risk data from discovery.

    Uses *_SERVICE_RISK_KB* to produce human-readable service names, risk
    narratives, attack examples and business-impact text for each related
    port.  Falls back to a generic message when a port is not in the KB.

    Mutates *findings_rows* in-place (adds ``related_ports`` list).
    """
    for f in findings_rows:
        f["related_ports"] = []
        scan_obj = scans_map.get(f.get("scan_id"))
        if not scan_obj:
            continue
        profile = device_profiles.get(scan_obj.target_id)
        if not profile or not profile.get("has_data"):
            continue

        open_port_numbers = {p.get("port") for p in profile.get("open_ports", []) if p.get("port")}
        if not open_port_numbers:
            continue

        # Build a searchable text blob from rule title + description (lowercase)
        text_blob = (f.get("rule_title", "") + " " + f.get("description", "")).lower()

        matched: list[dict] = []
        for keyword, ports in _SERVICE_PORT_MAP.items():
            if keyword in text_blob:
                for port_num in ports & open_port_numbers:
                    # Raw Nmap service name
                    svc_raw = next(
                        (p.get("service", str(port_num)) for p in profile["open_ports"] if p.get("port") == port_num),
                        str(port_num),
                    )
                    # Enrich from _SERVICE_RISK_KB
                    kb = _SERVICE_RISK_KB.get(port_num)
                    if kb:
                        matched.append({
                            "port": port_num,
                            "service_raw": svc_raw,
                            "service_name": kb["name"],
                            "service_purpose": kb["purpose"],
                            "matched_keyword": keyword,
                            "risk_narrative": kb["risk_when_rule_fails"],
                            "attack_examples": kb.get("attack_examples", ""),
                            "business_impact": kb.get("business_impact", ""),
                        })
                    else:
                        # Fallback: port not in KB — still show it, but with generic text
                        matched.append({
                            "port": port_num,
                            "service_raw": svc_raw,
                            "service_name": svc_raw,
                            "service_purpose": f"Network service on port {port_num}",
                            "matched_keyword": keyword,
                            "risk_narrative": f"This service ({svc_raw}) was found open on the target. The combination of an open port and a failed security control increases the exploitability of this finding.",
                            "attack_examples": "",
                            "business_impact": "",
                        })

        # Deduplicate by port
        seen_ports: set[int] = set()
        for m in matched:
            if m["port"] not in seen_ports:
                f["related_ports"].append(m)
                seen_ports.add(m["port"])


# ---------------------------------------------------------------------------
# Data aggregation
# ---------------------------------------------------------------------------

def aggregate_report_data(
    scope: str,
    scope_id: int | None,
    scan_ids: list[int] | None,
    db: Session,
    excluded_rule_ids: list[int] | None = None,
    severity_filter: list[str] | None = None,
) -> dict:
    """Query the database and return a normalised report data dict."""
    scans: list[Scan] = []
    _excluded = set(excluded_rule_ids) if excluded_rule_ids else set()
    _sev_filter = set(s.lower() for s in severity_filter) if severity_filter else None

    if scope == "scan":
        scan = (
            db.query(Scan)
            .options(joinedload(Scan.target).joinedload(Target.client))
            .filter(Scan.id == scope_id)
            .first()
        )
        if scan:
            scans = [scan]

    elif scope == "target":
        target = (
            db.query(Target)
            .options(joinedload(Target.client))
            .filter(Target.id == scope_id)
            .first()
        )
        if target:
            scans = db.query(Scan).filter(Scan.target_id == target.id).all()

    elif scope == "mission":
        mission = (
            db.query(Mission)
            .options(joinedload(Mission.client))
            .filter(Mission.id == scope_id)
            .first()
        )
        if mission:
            # Scans now have direct mission_id
            scans = db.query(Scan).filter(Scan.mission_id == mission.id).all()

    elif scope == "custom":
        if scan_ids:
            scans = (
                db.query(Scan)
                .options(joinedload(Scan.target).joinedload(Target.client))
                .filter(Scan.id.in_(scan_ids))
                .all()
            )

    # Determine client / mission context from the first available scan
    client_name = ""
    mission_name = ""
    is_finalized = False
    if scans:
        first_scan = scans[0]
        # Try mission from scan first (direct link)
        if first_scan.mission_id:
            mission_obj = db.query(Mission).filter(Mission.id == first_scan.mission_id).first()
            if mission_obj:
                mission_name = mission_obj.name or ""
                is_finalized = bool(mission_obj.is_locked)
                client_obj = db.query(Client).filter(Client.id == mission_obj.client_id).first()
                if client_obj:
                    client_name = client_obj.name or ""
        # Fallback: get client from target
        if not client_name:
            target_obj = db.query(Target).filter(Target.id == first_scan.target_id).first()
            if target_obj:
                client_obj = db.query(Client).filter(Client.id == target_obj.client_id).first()
                if client_obj:
                    client_name = client_obj.name or ""

    # Build per-target structure (batch-loaded)
    targets_map: dict[int, dict] = {}
    # Pre-load all targets and benchmarks for scans
    _target_ids = {s.target_id for s in scans}
    _targets_bulk = {t.id: t for t in db.query(Target).filter(Target.id.in_(_target_ids)).all()} if _target_ids else {}
    _bench_ids = {s.benchmark_id for s in scans if s.benchmark_id}
    _bench_bulk = {b.id: b for b in db.query(Benchmark).filter(Benchmark.id.in_(_bench_ids)).all()} if _bench_ids else {}

    for scan in scans:
        target_obj = _targets_bulk.get(scan.target_id)
        if not target_obj:
            continue
        if target_obj.id not in targets_map:
            # Parse os_details (may be JSON from get_system_info or a plain string)
            _os_raw = target_obj.os_details or ""
            _os_parsed = _parse_os_details(_os_raw)
            _os_display = _os_parsed.get("os", _os_raw) or ""
            if _os_parsed.get("os_version") and _os_parsed["os_version"] not in _os_display:
                _os_display = f"{_os_display} ({_os_parsed['os_version']})"
            if _os_parsed.get("architecture") and _os_parsed["architecture"] not in _os_display:
                _os_display = f"{_os_display} [{_os_parsed['architecture']}]"
            targets_map[target_obj.id] = {
                "hostname": target_obj.hostname or _os_parsed.get("hostname", ""),
                "ip_address": target_obj.ip_address or _os_parsed.get("ip", ""),
                "target_type": target_obj.target_type or "",
                "os_details": _os_display.strip(),
                "scans": [],
            }
        benchmark = _bench_bulk.get(scan.benchmark_id)
        targets_map[target_obj.id]["scans"].append({
            "id": scan.id,
            "benchmark_name": benchmark.name if benchmark else "",
            "scan_mode": scan.scan_mode or "",
            "compliance_percentage": scan.compliance_percentage,
            "passed": scan.passed or 0,
            "failed": scan.failed or 0,
            "errors": scan.errors or 0,
            "completed_at": str(scan.completed_at) if scan.completed_at else "",
        })

    # Collect findings
    scan_id_list = [s.id for s in scans]
    findings_rows: list[dict] = []
    total_passed = 0
    total_failed = 0
    total_errors = 0
    total_rules = 0
    severity_stats: dict[str, dict[str, int]] = {}

    if scan_id_list:
        db_findings = (
            db.query(Finding)
            .filter(Finding.scan_id.in_(scan_id_list))
            .all()
        )

        # ── Batch-load related entities to avoid N+1 queries ──
        rule_ids = {f.rule_id for f in db_findings if f.rule_id}
        rules_map: dict[int, Rule] = {}
        if rule_ids:
            for r in db.query(Rule).filter(Rule.id.in_(rule_ids)).all():
                rules_map[r.id] = r

        scans_map: dict[int, Scan] = {s.id: s for s in scans}
        target_ids = {s.target_id for s in scans if s.target_id}
        targets_lookup: dict[int, Target] = {}
        if target_ids:
            for t in db.query(Target).filter(Target.id.in_(target_ids)).all():
                targets_lookup[t.id] = t

        benchmark_ids = {s.benchmark_id for s in scans if s.benchmark_id}
        benchmarks_lookup: dict[int, Benchmark] = {}
        if benchmark_ids:
            for b in db.query(Benchmark).filter(Benchmark.id.in_(benchmark_ids)).all():
                benchmarks_lookup[b.id] = b

        for f in db_findings:
            # Skip excluded rules
            if _excluded and f.rule_id in _excluded:
                continue
            rule = rules_map.get(f.rule_id)
            scan_obj = scans_map.get(f.scan_id)
            target_obj = targets_lookup.get(scan_obj.target_id) if scan_obj else None
            benchmark = benchmarks_lookup.get(scan_obj.benchmark_id) if scan_obj else None

            sev = (f.severity or (rule.severity if rule else "medium") or "medium").lower()
            status = (f.status or "").upper()

            # Skip findings outside the severity filter
            if _sev_filter and sev not in _sev_filter:
                continue

            findings_rows.append({
                "_rule_id": f.rule_id,
                "scan_id": f.scan_id,
                "target_hostname": target_obj.hostname if target_obj else "",
                "target_type": target_obj.target_type if target_obj else "",
                "benchmark_name": benchmark.name if benchmark else "",
                "section_number": rule.section_number if rule else "",
                "rule_title": normalize_unicode(rule.title) if rule else "",
                "description": normalize_unicode(rule.description) if rule else "",
                "rationale": normalize_unicode(rule.rationale) if rule else "",
                "default_value": normalize_unicode(rule.default_value) if rule else "",
                "severity": sev,
                "status": status,
                "actual_output": normalize_unicode(f.actual_output or ""),
                "expected_output": normalize_unicode(f.expected_output or ""),
                "remediation": normalize_unicode(rule.remediation_description_raw) if rule else "",
                "evaluation_explanation": normalize_unicode(f.evaluation_explanation or ""),
                "ai_advice": normalize_unicode(f.ai_advice or ""),
                "auditor_notes": normalize_unicode(f.auditor_notes or ""),
                "auditor_override": normalize_unicode(f.auditor_override or ""),
                "security_themes": _safe_json_loads(rule.security_themes_json, []) if rule and rule.security_themes_json else [],
                "framework_mappings": _safe_json_loads(rule.framework_mappings, {}) if rule and rule.framework_mappings else {},
            })

            total_rules += 1
            if status == "PASS":
                total_passed += 1
            elif status == "FAIL":
                total_failed += 1
            elif status == "ERROR":
                total_errors += 1

            if sev not in severity_stats:
                severity_stats[sev] = {"total": 0, "passed": 0, "failed": 0}
            severity_stats[sev]["total"] += 1
            if status == "PASS":
                severity_stats[sev]["passed"] += 1
            elif status == "FAIL":
                severity_stats[sev]["failed"] += 1

    overall_compliance = round((total_passed / total_rules) * 100, 1) if total_rules > 0 else 0.0

    # Date range
    completed_dates = [s.completed_at for s in scans if s.completed_at]
    if completed_dates:
        earliest = min(completed_dates)
        latest = max(completed_dates)
        date_range = f"{earliest.strftime('%Y-%m-%d')} to {latest.strftime('%Y-%m-%d')}"
    else:
        date_range = "N/A"

    # Per-target compliance summary
    by_target: list[dict] = []
    for _tid, tdata in targets_map.items():
        t_p = sum(s["passed"] for s in tdata["scans"])
        t_f = sum(s["failed"] for s in tdata["scans"])
        t_e = sum(s["errors"] for s in tdata["scans"])
        t_total = t_p + t_f + t_e
        by_target.append({
            "hostname": tdata["hostname"],
            "ip_address": tdata["ip_address"],
            "compliance": round((t_p / t_total) * 100, 1) if t_total > 0 else 0,
            "passed": t_p, "failed": t_f, "errors": t_e, "total": t_total,
        })

    # Category-level stats (group by 1st section number)
    by_category: dict[str, dict] = {}
    for f in findings_rows:
        sec = f.get("section_number", "")
        cat = sec.split(".")[0] if sec else "Other"
        if cat not in by_category:
            by_category[cat] = {"total": 0, "passed": 0, "failed": 0, "label": f"Section {cat}"}
        by_category[cat]["total"] += 1
        if f["status"] == "PASS":
            by_category[cat]["passed"] += 1
        elif f["status"] == "FAIL":
            by_category[cat]["failed"] += 1

    # Enrich category labels with real CIS section names
    for cat_key, info in by_category.items():
        info["label"] = _resolve_section_name(cat_key, findings_rows)

    # ── Sort findings by section number for consistent ordering ──
    def _section_sort_key(f):
        """Sort by numeric section parts (e.g., 1.1.1 → (1, 1, 1))."""
        parts = []
        for p in (f.get("section_number") or "").split("."):
            try:
                parts.append(int(p))
            except ValueError:
                parts.append(999)
        return parts

    findings_rows.sort(key=_section_sort_key)

    # ── Smart output truncation for display ──
    for f in findings_rows:
        actual_t = smart_truncate(f["actual_output"], max_chars=400)
        expected_t = smart_truncate(f["expected_output"], max_chars=400)
        f["actual_output_display"] = actual_t["text"]
        f["actual_output_truncated"] = actual_t["truncated"]
        f["actual_output_length"] = actual_t["original_length"]
        f["actual_output_lines_hidden"] = actual_t["lines_hidden"]
        f["expected_output_display"] = expected_t["text"]
        f["expected_output_truncated"] = expected_t["truncated"]
        f["expected_output_length"] = expected_t["original_length"]

    # ── False-positive detection ──
    findings_rows, fp_summary = enrich_findings_with_fp(findings_rows)

    # ── Device profiles from DiscoveryCache ──
    device_profiles = _build_device_profiles(targets_map, _targets_bulk, db)

    # ── Enrich bare port entries with service names from Risk KB ──
    _enrich_port_entries(device_profiles)

    # ── Link findings to open services (informational only) ──
    _scans_map_for_link = {s.id: s for s in scans}
    _link_findings_to_services(findings_rows, device_profiles, _scans_map_for_link)

    # Derive a top-level benchmark_name from the first scan's benchmark
    _first_bench = next(
        (b for b in _bench_bulk.values()), None
    )

    # Build audit_info for the appendix (benchmark metadata)
    audit_info = {}
    if _first_bench:
        audit_info = {
            "benchmark_name": _first_bench.name or "CIS Benchmark",
            "benchmark_version": _first_bench.version or "N/A",
            "profile_level": _first_bench.platform or "N/A",
        }

    return {
        "title": "",
        "client_name": client_name,
        "mission_name": mission_name,
        "benchmark_name": _first_bench.name if _first_bench else "CIS Benchmark",
        "date_range": date_range,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "targets": list(targets_map.values()),
        "findings": findings_rows,
        "summary": {
            "total_rules": total_rules,
            "passed": total_passed,
            "failed": total_failed,
            "errors": total_errors,
            "overall_compliance": overall_compliance,
            "by_severity": severity_stats,
        },
        "scans": [
            {
                "id": s.id,
                "target_id": s.target_id,
                "benchmark_id": s.benchmark_id,
                "status": s.status,
                "compliance_percentage": s.compliance_percentage,
                "completed_at": str(s.completed_at) if s.completed_at else "",
            }
            for s in scans
        ],
        "by_target": by_target,
        "by_category": by_category,
        "categories_enriched": _build_enriched_categories({"by_category": by_category, "findings": findings_rows}),
        "fp_summary": fp_summary,
        "device_profiles": device_profiles,
        "audit_info": audit_info,
        "ai_summary": None,
        "is_finalized": is_finalized,
    }


# ---------------------------------------------------------------------------
# Helper: build grouped findings for report builder
# ---------------------------------------------------------------------------

_SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}


def _enrich_group(name: str, finds: list[dict], summary_text: str) -> dict:
    """Build a fully-enriched group dict with severity breakdown, compliance, sorted findings."""
    # Sort findings: severity order (critical→low), then FAIL before PASS
    finds_sorted = sorted(
        finds,
        key=lambda f: (_SEV_ORDER.get(f.get("severity", "medium"), 5), 0 if f.get("status") == "FAIL" else 1),
    )

    pass_count = sum(1 for f in finds if f.get("status") == "PASS")
    fail_count = sum(1 for f in finds if f.get("status") == "FAIL")
    error_count = sum(1 for f in finds if f.get("status") == "ERROR")
    total = len(finds)
    compliance_pct = round((pass_count / total) * 100, 1) if total > 0 else 0.0

    # Severity breakdown (failed only — for risk heatmap)
    sev_counts: dict[str, int] = {}
    sev_detail: dict[str, dict[str, int]] = {}
    for sev in ("critical", "high", "medium", "low", "informational"):
        s_total = sum(1 for f in finds if f.get("severity") == sev)
        s_pass = sum(1 for f in finds if f.get("severity") == sev and f.get("status") == "PASS")
        s_fail = sum(1 for f in finds if f.get("severity") == sev and f.get("status") == "FAIL")
        sev_counts[sev] = s_fail  # used in heatmap (failed count)
        if s_total > 0:
            sev_detail[sev] = {"total": s_total, "passed": s_pass, "failed": s_fail}

    # Exposed ports summary (from port-risk engine)
    exposed_ports: list[dict] = []
    _seen_ep: set[int] = set()
    for f in finds:
        for rp in f.get("related_ports", []):
            if rp["port"] not in _seen_ep:
                exposed_ports.append({
                    "port": rp["port"],
                    "service_name": rp.get("service_name", rp.get("service_raw", str(rp["port"]))),
                })
                _seen_ep.add(rp["port"])
    exposed_ports_summary = ""
    if exposed_ports:
        names = ", ".join(ep["service_name"] for ep in exposed_ports)
        exposed_ports_summary = f"{len(exposed_ports)} exposed service(s): {names}"

    return {
        "name": name,
        "findings": finds_sorted,
        "summary": summary_text,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "error_count": error_count,
        "total_count": total,
        "compliance_pct": compliance_pct,
        "sev_counts": sev_counts,
        "sev_detail": sev_detail,
        "exposed_ports": exposed_ports,
        "exposed_ports_summary": exposed_ports_summary,
    }


# ---------------------------------------------------------------------------
# Phase 4 — Business Impact Section (templated, no AI)
# ---------------------------------------------------------------------------

_THEME_DISPLAY = {
    "remote_access": "Remote Access Exposure",
    "authentication": "Authentication & Password Weaknesses",
    "audit_logging": "Audit Trail Gaps",
    "network_security": "Network Security & Firewall",
    "data_protection": "Data Protection & Encryption",
    "privilege_escalation": "Privilege Escalation Risks",
    "system_integrity": "System Integrity",
}

_THEME_TEMPLATES: dict[str, dict[str, str]] = {
    "remote_access": {
        "summary": "{count} remote access security control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "Combined with open {port_list} port(s), this creates a direct path for unauthorized remote access to {target_desc}.",
        "port_context_without": "While no corresponding remote access ports were detected as open, the misconfiguration should still be remediated to prevent future exposure.",
        "business_risk": "Unauthorized remote access to {target_desc}. Attackers could deploy ransomware, exfiltrate sensitive data, or establish persistent access to the corporate network.",
    },
    "authentication": {
        "summary": "{count} authentication/password control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "With {port_list} exposed, credential-based attacks (brute-force, password spraying) can be conducted remotely.",
        "port_context_without": "Even without direct network exposure, weak authentication policies enable insider threats and credential-based attacks.",
        "business_risk": "Weak authentication policies allow unauthorized account access. This may lead to data breaches, regulatory non-compliance ({reg_refs}), and compromise of privileged accounts.",
    },
    "audit_logging": {
        "summary": "{count} audit/logging control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "With {port_list} exposed, attackers can operate undetected on systems lacking proper audit trails.",
        "port_context_without": "Insufficient logging reduces the ability to detect and investigate security incidents.",
        "business_risk": "Loss of audit trail integrity impairs incident detection. Regulatory frameworks ({reg_refs}) require comprehensive logging for compliance. A breach without logs may increase liability.",
    },
    "network_security": {
        "summary": "{count} network/firewall control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "Combined with open {port_list} port(s), the attack surface is significantly expanded.",
        "port_context_without": "Misconfigured network controls may allow lateral movement within the internal network.",
        "business_risk": "Weakened network defenses expose {target_desc} to lateral movement and network-based attacks. Ransomware and worm propagation are the primary business risks.",
    },
    "data_protection": {
        "summary": "{count} data protection/encryption control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "With {port_list} exposed, unencrypted data in transit or at rest is vulnerable to interception.",
        "port_context_without": "Unencrypted storage or weak TLS configurations expose data to physical theft and insider threats.",
        "business_risk": "Sensitive data exposure risk — customer PII, financial records, or credentials may be compromised. Regulatory penalties ({reg_refs}) for data protection failures can be substantial.",
    },
    "privilege_escalation": {
        "summary": "{count} privilege/access control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "With {port_list} exposed, attackers who gain initial access can quickly escalate privileges.",
        "port_context_without": "Overly permissive privileges allow insider threats and post-compromise escalation.",
        "business_risk": "Privilege escalation enables attackers to gain administrative control over {target_desc}. This is a precondition for data exfiltration, domain compromise, and persistent backdoor installation.",
    },
    "system_integrity": {
        "summary": "{count} system integrity control(s) are failing across {target_count} system(s). {port_context}",
        "port_context_with": "With {port_list} exposed, attackers can exploit unpatched vulnerabilities or tamper with boot configuration.",
        "port_context_without": "Missing integrity controls (updates, code signing, secure boot) increase susceptibility to malware and rootkits.",
        "business_risk": "Compromised system integrity undermines trust in the computing environment. Malware persistence, ransomware, and supply-chain attacks become feasible.",
    },
}

# Keywords to auto-detect themes from rule titles/descriptions when security_themes_json is empty
_THEME_KEYWORDS: dict[str, list[str]] = {
    "remote_access": ["rdp", "remote desktop", "ssh", "vnc", "winrm", "remote access", "remote management", "terminal service"],
    "authentication": ["password", "account lockout", "credential", "logon", "kerberos", "authentication", "smart card", "brute"],
    "audit_logging": ["audit", "event log", "logging", "log size", "log retention", "security log"],
    "network_security": ["firewall", "smb", "netbios", "network access", "icmp", "ip source routing", "tcp/ip", "wifi", "wlan"],
    "data_protection": ["bitlocker", "encrypt", "tls", "ssl", "certificate", "data recovery", "drive encryption"],
    "privilege_escalation": ["uac", "admin approval", "privilege", "elevation", "run as", "installer detection", "user rights", "access control"],
    "system_integrity": ["update", "windows update", "code signing", "device guard", "secure boot", "wdac", "applocker", "defender"],
}


def _detect_themes(finding: dict) -> list[str]:
    """Auto-detect themes from rule title and description when security_themes_json is empty."""
    text = f"{finding.get('rule_title', '')} {finding.get('description', '')}".lower()
    themes = []
    for theme, keywords in _THEME_KEYWORDS.items():
        if any(kw in text for kw in keywords):
            themes.append(theme)
    return themes


def _generate_business_impact(findings: list[dict]) -> list[dict]:
    """Generate business impact groups from failed critical/high findings, clustered by theme."""
    # Only consider failed findings with critical or high severity
    relevant = [f for f in findings if f.get("status") == "FAIL" and f.get("severity") in ("critical", "high")]
    if not relevant:
        return []

    # Cluster findings by theme
    theme_buckets: dict[str, list[dict]] = {}
    for f in relevant:
        themes = f.get("security_themes", []) or _detect_themes(f)
        if not themes:
            themes = ["system_integrity"]  # fallback bucket
        for theme in themes:
            theme_buckets.setdefault(theme, []).append(f)

    # Build business impact groups
    groups = []
    for theme, theme_findings in theme_buckets.items():
        templates = _THEME_TEMPLATES.get(theme, _THEME_TEMPLATES["system_integrity"])
        display_name = _THEME_DISPLAY.get(theme, theme.replace("_", " ").title())

        # Determine worst severity
        worst_sev = "high"
        if any(f.get("severity") == "critical" for f in theme_findings):
            worst_sev = "critical"

        # Collect affected targets
        targets_set: set[str] = set()
        target_types: set[str] = set()
        for f in theme_findings:
            if f.get("target_hostname"):
                targets_set.add(f["target_hostname"])
            if f.get("target_type"):
                target_types.add(f["target_type"])

        target_desc = "critical infrastructure"
        if target_types:
            parts = []
            for tt in sorted(target_types):
                count = sum(1 for f in theme_findings if f.get("target_type") == tt)
                parts.append(f"{tt}s" if count > 1 else f"a {tt}")
            target_desc = ", ".join(parts)

        # Collect exposed ports from related_ports
        exposed_ports: list[dict] = []
        _seen_port_target: set[tuple] = set()
        for f in theme_findings:
            hostname = f.get("target_hostname", "")
            for rp in f.get("related_ports", []):
                key = (rp["port"], hostname)
                if key not in _seen_port_target:
                    exposed_ports.append({
                        "port": rp["port"],
                        "service_name": rp.get("service_name", str(rp["port"])),
                        "target": hostname,
                    })
                    _seen_port_target.add(key)

        # Build port list string
        port_list = ", ".join(f"{ep['service_name']} ({ep['port']})" for ep in exposed_ports[:8])

        # Build port context
        if exposed_ports:
            port_context = templates["port_context_with"].format(
                port_list=port_list, target_desc=target_desc
            )
        else:
            port_context = templates["port_context_without"]

        # Collect regulatory references
        reg_refs: list[str] = []
        for f in theme_findings:
            fm = f.get("framework_mappings", {})
            if isinstance(fm, dict):
                for framework, refs in fm.items():
                    if isinstance(refs, list):
                        for ref in refs:
                            tag = f"{framework.replace('_', ' ')}: {ref}"
                            if tag not in reg_refs:
                                reg_refs.append(tag)
        reg_refs = reg_refs[:6]  # cap for readability

        reg_refs_str = ", ".join(reg_refs) if reg_refs else "applicable security frameworks"

        # Build summary
        summary = templates["summary"].format(
            count=len(theme_findings),
            target_count=len(targets_set) or 1,
            port_context=port_context,
        )

        # Build business risk
        business_risk = templates["business_risk"].format(
            target_desc=target_desc,
            reg_refs=reg_refs_str,
        )

        groups.append({
            "theme": display_name,
            "theme_key": theme,
            "severity": worst_sev,
            "finding_count": len(theme_findings),
            "affected_targets": sorted(targets_set),
            "exposed_ports": exposed_ports,
            "summary": summary,
            "business_risk": business_risk,
            "regulatory_refs": reg_refs,
            "findings": theme_findings,
        })

    # Sort by severity (critical first), then by finding count descending
    groups.sort(key=lambda g: (0 if g["severity"] == "critical" else 1, -g["finding_count"]))
    return groups


def _build_grouped_findings(data: dict, findings: list[dict]) -> list[dict] | None:
    """Build grouped findings from builder_groups in data.

    Returns a list of enriched group dicts or None if no builder groups are set.
    Each group contains: name, findings (sorted by severity), summary, pass/fail/error counts,
    compliance_pct, sev_counts (failed by severity), sev_detail (per-severity breakdown).
    """
    builder_groups = data.get("builder_groups")
    if not builder_groups:
        return None

    group_summaries = data.get("group_summaries") or {}

    grouped = []
    for bg in builder_groups:
        rule_id_set = set(bg["rule_ids"])
        group_finds = [f for f in findings if f.get("_rule_id") in rule_id_set]
        grouped.append(_enrich_group(bg["name"], group_finds, group_summaries.get(bg["name"], "")))

    # Add ungrouped findings
    grouped_ids = set()
    for bg in builder_groups:
        grouped_ids.update(bg["rule_ids"])
    ungrouped = [f for f in findings if f.get("_rule_id") not in grouped_ids]
    if ungrouped:
        grouped.append(_enrich_group("Other", ungrouped, group_summaries.get("Other", "")))

    return grouped


# ---------------------------------------------------------------------------
# Shared chart generation (used by both PDF and HTML)
# ---------------------------------------------------------------------------

def _generate_all_charts(data: dict, grouped_findings: list[dict] | None) -> dict:
    """Generate all SVG charts and return as a dict of chart_name → SVG string.

    This avoids duplicating chart generation logic between PDF and HTML exporters.
    """
    from backend.core.chart_helpers import (
        generate_donut_svg,
        generate_hbar_svg,
        generate_risk_heatmap_svg,
        generate_mini_donut_svg,
        generate_stacked_hbar_svg,
        generate_fp_gauge_svg,
        generate_treemap_svg,
        generate_radar_svg,
        generate_waterfall_svg,
    )

    summary = data["summary"]
    charts: dict[str, str] = {}

    # 1. Donut: results distribution
    charts["chart_donut"] = generate_donut_svg([
        {"label": "Passed", "value": summary["passed"], "color": "#22c55e"},
        {"label": "Failed", "value": summary["failed"], "color": "#ef4444"},
        {"label": "Errors", "value": summary["errors"], "color": "#8b5cf6"},
    ], title="Results Distribution", size=220)

    # 2. Severity compliance bars
    sev_items = []
    for sev in ("critical", "high", "medium", "low", "informational"):
        info = summary["by_severity"].get(sev, {"total": 0, "passed": 0})
        comp = round((info["passed"] / info["total"]) * 100, 1) if info["total"] > 0 else 0
        colors = {"critical": "#dc2626", "high": "#ea580c", "medium": "#d97706", "low": "#2563eb", "informational": "#6b7280"}
        sev_items.append({"label": sev.capitalize(), "value": comp, "color": colors[sev]})
    charts["chart_severity"] = generate_hbar_svg(sev_items, title="Compliance by Severity", width=340)

    # 3. Per-target compliance bars
    target_items = []
    for t in data.get("by_target", []):
        c = t["compliance"]
        color = "#22c55e" if c >= 80 else "#f59e0b" if c >= 50 else "#ef4444"
        target_items.append({"label": t["hostname"] or t["ip_address"], "value": c, "color": color})
    charts["chart_targets"] = generate_hbar_svg(target_items, title="Compliance by Target", width=340) if len(target_items) > 1 else ""

    # 4. Category stacked bar
    by_cat = data.get("by_category", {})
    cat_items = [
        {"label": info.get("label", f"Section {k}"), "passed": info.get("passed", 0), "failed": info.get("failed", 0)}
        for k, info in sorted(by_cat.items(), key=lambda x: (not x[0].isdigit(), x[0]))
    ]
    charts["chart_categories"] = generate_stacked_hbar_svg(cat_items, title="Findings by Category", width=450) if cat_items else ""

    # 5. Grouped charts & risk heatmap
    charts["chart_risk_heatmap"] = ""
    charts["chart_group_compliance"] = ""
    if grouped_findings:
        hm_groups = [{"name": g["name"], "sev_counts": g["sev_counts"]} for g in grouped_findings]
        charts["chart_risk_heatmap"] = generate_risk_heatmap_svg(hm_groups)

        gc_items = [
            {"label": g["name"], "passed": g["pass_count"], "failed": g["fail_count"]}
            for g in grouped_findings
        ]
        charts["chart_group_compliance"] = generate_stacked_hbar_svg(gc_items, title="Compliance by Group", width=450)

        for g in grouped_findings:
            g["mini_donut"] = generate_mini_donut_svg(g["pass_count"], g["fail_count"], g["error_count"])

    # 6. False-positive gauge
    fp_summary = data.get("fp_summary", {})
    charts["chart_fp_gauge"] = ""
    if fp_summary.get("total_suspects", 0) > 0:
        charts["chart_fp_gauge"] = generate_fp_gauge_svg(
            fp_summary.get("high_confidence", 0),
            fp_summary.get("medium_confidence", 0),
            fp_summary.get("low_confidence", 0),
            summary["failed"],
        )

    # 7. Category treemap
    cat_treemap_items = [
        {"label": info.get("label", f"Section {k}"), "total": info["total"], "passed": info["passed"], "failed": info["failed"]}
        for k, info in sorted(by_cat.items(), key=lambda x: (not x[0].isdigit(), x[0]))
    ]
    charts["chart_treemap"] = generate_treemap_svg(cat_treemap_items, title="Category Compliance Map") if cat_treemap_items else ""

    # 8. Compliance radar
    radar_cats = [
        {"label": info.get("label", f"Section {k}"),
         "compliance": round((info["passed"] / info["total"]) * 100, 1) if info["total"] > 0 else 0}
        for k, info in sorted(by_cat.items(), key=lambda x: (not x[0].isdigit(), x[0]))
    ]
    charts["chart_radar"] = generate_radar_svg(radar_cats, title="Compliance by Category") if len(radar_cats) >= 3 else ""

    # 9. Compliance waterfall
    charts["chart_waterfall"] = generate_waterfall_svg(
        cat_treemap_items, summary["total_rules"], summary["passed"],
        title="Compliance Waterfall - Failures by Category",
    ) if cat_treemap_items else ""

    return charts


# ---------------------------------------------------------------------------
# CIS Benchmark section name mapping
# ---------------------------------------------------------------------------

# Maps top-level section numbers to descriptive CIS benchmark section names.
# Covers CIS Microsoft Windows 11 and common CIS benchmarks.
_CIS_SECTION_NAMES: dict[str, str] = {
    "1": "Account Policies",
    "2": "Local Policies",
    "3": "Event Log",
    "4": "Restricted Groups",
    "5": "System Services",
    "6": "Registry",
    "7": "File System",
    "8": "Wired Network Policies",
    "9": "Windows Firewall",
    "10": "Network List Manager",
    "11": "Wireless Network",
    "12": "Public Key Policies",
    "13": "Software Restriction",
    "14": "Network Access Protection",
    "15": "Application Control",
    "16": "IP Security Policies",
    "17": "Advanced Audit Policy",
    "18": "Administrative Templates (Computer)",
    "19": "Administrative Templates (User)",
    "20": "File & Storage",
    "21": "Print & Document",
    "22": "Remote Services",
}


def _resolve_section_name(section_num: str, findings: list[dict]) -> str:
    """Try to resolve a meaningful name for a top-level section number.

    1. Check the CIS section name map.
    2. Fall back to the first common word in rule titles for that section.
    3. Default to 'Section N'.
    """
    if section_num in _CIS_SECTION_NAMES:
        return _CIS_SECTION_NAMES[section_num]

    # Attempt to derive from rule titles in that section
    titles = [f["rule_title"] for f in findings
              if (f.get("section_number") or "").split(".")[0] == section_num and f.get("rule_title")]
    if titles:
        # Find most common leading word(s)
        first_title = titles[0]
        # Use "Ensure" prefix removal for cleaner label
        clean = first_title.replace("Ensure ", "").replace("(L1) ", "").replace("(L2) ", "")
        if len(clean) > 40:
            clean = clean[:38] + ".."
        return clean

    return f"Section {section_num}"


def _build_enriched_categories(data: dict) -> list[dict]:
    """Build enriched category list with resolved CIS section names and compliance bars."""
    by_category = data.get("by_category", {})
    findings = data.get("findings", [])
    enriched = []
    for cat_key in sorted(by_category.keys(), key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else 0, x)):
        info = by_category[cat_key]
        comp = round((info["passed"] / info["total"]) * 100, 1) if info["total"] > 0 else 0.0
        enriched.append({
            "key": cat_key,
            "label": _resolve_section_name(cat_key, findings),
            "total": info["total"],
            "passed": info["passed"],
            "failed": info["failed"],
            "compliance": comp,
        })
    return enriched


def _compute_risk_score(data: dict) -> dict:
    """Compute an overall risk score with letter grade (A-F) for the audit."""
    summary = data.get("summary", {})
    compliance = summary.get("overall_compliance", 0)
    by_sev = summary.get("by_severity", {})

    # Weighted severity penalty: critical failures weigh much more
    crit_fails = by_sev.get("critical", {}).get("failed", 0)
    high_fails = by_sev.get("high", {}).get("failed", 0)
    med_fails = by_sev.get("medium", {}).get("failed", 0)
    low_fails = by_sev.get("low", {}).get("failed", 0)
    total_rules = summary.get("total_rules", 1) or 1

    # Weighted score: compliance % minus severity penalties
    penalty = (crit_fails * 5 + high_fails * 3 + med_fails * 1 + low_fails * 0.3) / total_rules * 10
    score = max(0, min(100, compliance - penalty))

    if score >= 90:
        grade, label, color = "A", "Excellent", "#16a34a"
    elif score >= 80:
        grade, label, color = "B", "Good", "#22c55e"
    elif score >= 70:
        grade, label, color = "C", "Acceptable", "#f59e0b"
    elif score >= 50:
        grade, label, color = "D", "Poor", "#ea580c"
    else:
        grade, label, color = "F", "Critical", "#dc2626"

    risk_level = "LOW" if score >= 80 else "MODERATE" if score >= 60 else "HIGH" if score >= 40 else "CRITICAL"

    return {
        "score": round(score, 1),
        "grade": grade,
        "label": label,
        "color": color,
        "risk_level": risk_level,
        "critical_fails": crit_fails,
        "high_fails": high_fails,
        "medium_fails": med_fails,
        "low_fails": low_fails,
    }


def _build_per_target_findings(data: dict) -> list[dict]:
    """Group findings by target hostname for per-target breakdown section."""
    findings = data.get("findings", [])
    targets: dict[str, dict] = {}
    for f in findings:
        host = f.get("target_hostname") or "Unknown"
        if host not in targets:
            targets[host] = {"hostname": host, "total": 0, "passed": 0, "failed": 0, "errors": 0,
                             "critical_fails": 0, "high_fails": 0, "findings": []}
        targets[host]["total"] += 1
        st = f.get("status", "")
        if st == "PASS":
            targets[host]["passed"] += 1
        elif st == "FAIL":
            targets[host]["failed"] += 1
            sev = f.get("severity", "")
            if sev == "critical":
                targets[host]["critical_fails"] += 1
            elif sev == "high":
                targets[host]["high_fails"] += 1
        elif st == "ERROR":
            targets[host]["errors"] += 1
        targets[host]["findings"].append(f)

    result = []
    for host, tdata in targets.items():
        comp = round((tdata["passed"] / tdata["total"]) * 100, 1) if tdata["total"] > 0 else 0.0
        tdata["compliance"] = comp
        # Top 5 failures per target
        tdata["top_failures"] = [f for f in tdata["findings"] if f["status"] == "FAIL"][:5]
        result.append(tdata)
    return sorted(result, key=lambda t: t["compliance"])


# ---------------------------------------------------------------------------
# White-label settings loader
# ---------------------------------------------------------------------------

def _load_whitelabel_settings(db: Session) -> dict[str, str]:
    """Load company_name, company_logo_base64, and auditor_name from AppSettings."""
    from backend.models.app_settings import AppSettings
    keys = ("company_name", "company_logo_base64", "auditor_name")
    rows = db.query(AppSettings).filter(AppSettings.key.in_(keys)).all()
    return {row.key: (row.value or "") for row in rows}


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def generate_pdf_report(data: dict, include_passed: bool, db: Session) -> bytes:
    """Render an HTML template with Jinja2 and convert to PDF via WeasyPrint."""
    from weasyprint import HTML  # imported here to isolate heavy dep

    env = Environment(loader=FileSystemLoader(TEMPLATES_DIR), autoescape=True)
    template = env.get_template("report.html.j2")

    if not include_passed:
        data = {**data, "findings": [f for f in data["findings"] if f["status"] != "PASS"]}

    # ── White-label settings ──
    _wl = _load_whitelabel_settings(db)
    data["company_name"] = _wl.get("company_name", "AuditForge")
    data["company_logo_base64"] = _wl.get("company_logo_base64", "")
    data["auditor_name"] = _wl.get("auditor_name", "")
    data["project_logo_base64"] = _PROJECT_LOGO_B64

    # Phase 2: grouped findings & builder metadata for template
    grouped_findings = _build_grouped_findings(data, data["findings"])
    data["grouped_findings"] = grouped_findings
    data["section_toggles"] = data.get("sections") or {}
    data["group_summaries"] = data.get("group_summaries") or {}
    data["audience"] = data.get("audience", "technical")
    data["has_builder"] = bool(data.get("builder_groups"))

    # Generate all SVG charts (shared logic)
    charts = _generate_all_charts(data, grouped_findings)
    data.update(charts)
    data["fp_summary"] = data.get("fp_summary", {})

    # ── Convert AI summary from Markdown to HTML ──
    if data.get("ai_summary"):
        data["ai_summary_html"] = _md_to_html(data["ai_summary"])
    else:
        data["ai_summary_html"] = ""

    # categories_enriched already computed in aggregate_report_data()

    # ── Compute overall risk score (A-F letter grade) ──
    data["risk_score"] = _compute_risk_score(data)

    # ── Build per-target findings breakdown ──
    data["per_target_findings"] = _build_per_target_findings(data)

    # ── Build device profiles list for template ──
    dp = data.get("device_profiles", {})
    data["device_profiles_list"] = list(dp.values())
    data["has_device_profiles"] = len(data["device_profiles_list"]) > 0

    # Build top-N remediation items for the Recommendations section
    _sev_prio = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    failed_for_recs = [f for f in data["findings"] if f.get("status") == "FAIL"]
    failed_for_recs.sort(key=lambda f: (_sev_prio.get(f.get("severity", "medium"), 5),
                                         f.get("section_number", "")))
    data["recommendations"] = failed_for_recs  # template will use this for a dedicated section

    # Phase 4: Business Impact Section
    data["business_impact_groups"] = _generate_business_impact(data["findings"])

    html_string = template.render(**data, include_passed=include_passed)

    try:
        pdf_bytes = HTML(string=html_string).write_pdf()
    except Exception as exc:
        logger.exception("WeasyPrint PDF generation failed")
        raise RuntimeError(f"PDF generation failed: {exc}") from exc

    return pdf_bytes


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_excel_report(data: dict, include_passed: bool) -> bytes:
    """Create a multi-sheet Excel workbook and return as bytes."""
    wb = Workbook()

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active
    ws1.title = "Executive Summary"
    summary = data["summary"]
    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    label_font = Font(bold=True)

    ws1.append(["Audit Report - Executive Summary"])
    ws1["A1"].font = header_font
    ws1.append([])
    ws1.append(["Client", data.get("client_name", "")])
    ws1.append(["Mission", data.get("mission_name", "")])
    ws1.append(["Benchmark", data.get("benchmark_name", "CIS Benchmark")])
    _ai = data.get("audit_info", {})
    if _ai.get("benchmark_version"):
        ws1.append(["Benchmark Version", _ai["benchmark_version"]])
    if _ai.get("profile_level"):
        ws1.append(["Platform / Profile", _ai["profile_level"]])
    ws1.append(["Date Range", data.get("date_range", "")])
    ws1.append(["Generated", data.get("generated_at", "")])
    ws1.append([])
    ws1.append(["Metric", "Value"])
    _metric_row = ws1.max_row
    ws1.cell(row=_metric_row, column=1).font = label_font
    ws1.cell(row=_metric_row, column=2).font = label_font
    ws1.append(["Overall Compliance", f"{summary['overall_compliance']}%"])
    ws1.append(["Total Rules", summary["total_rules"]])
    ws1.append(["Passed", summary["passed"]])
    ws1.append(["Failed", summary["failed"]])
    ws1.append(["Errors", summary["errors"]])
    ws1.append([])
    ws1.append(["Severity", "Total", "Passed", "Failed"])
    _sev_header_row = ws1.max_row
    for _ci in range(1, 5):
        ws1.cell(row=_sev_header_row, column=_ci).font = label_font
    for sev in ("critical", "high", "medium", "low", "informational"):
        info = summary["by_severity"].get(sev, {"total": 0, "passed": 0, "failed": 0})
        row_idx = ws1.max_row + 1
        ws1.append([sev.capitalize(), info["total"], info["passed"], info["failed"]])
        fill = SEVERITY_FILLS.get(sev)
        if fill:
            ws1.cell(row=row_idx, column=1).fill = fill

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Findings ──
    ws2 = wb.create_sheet("Findings")
    columns = [
        "Scan ID", "Target", "Benchmark", "Section", "Rule Title",
        "Severity", "Status", "Description", "Actual Output", "Expected Output",
        "Remediation", "Evaluation", "AI Advice", "Auditor Notes", "Auditor Override",
        "Related Services",
    ]
    ws2.append(columns)
    for col_idx in range(1, len(columns) + 1):
        ws2.cell(row=1, column=col_idx).font = label_font

    def _clean(val):
        """Strip illegal XML/Excel control characters from a string value."""
        if isinstance(val, str):
            return _ILLEGAL_XLSX_RE.sub('', val)
        return val

    for f in data["findings"]:
        if not include_passed and f["status"] == "PASS":
            continue
        row_idx = ws2.max_row + 1
        ws2.append([
            f["scan_id"],
            _clean(f["target_hostname"]),
            _clean(f["benchmark_name"]),
            f["section_number"],
            _clean(f["rule_title"]),
            f["severity"],
            f["status"],
            _clean(f.get("description", "")),
            _clean(f["actual_output"]),
            _clean(f["expected_output"]),
            _clean(f["remediation"]),
            _clean(f.get("evaluation_explanation", "")),
            _clean(f.get("ai_advice", "")),
            _clean(f.get("auditor_notes", "")),
            _clean(f.get("auditor_override", "")),
            ", ".join(f"{rp.get('service_name', rp.get('service', ''))}:{rp.get('port', '')}" for rp in f.get("related_ports", [])) or "",
        ])
        # Color-code severity
        sev_fill = SEVERITY_FILLS.get(f["severity"])
        if sev_fill:
            ws2.cell(row=row_idx, column=6).fill = sev_fill
        # Color-code status
        if f["status"] == "PASS":
            ws2.cell(row=row_idx, column=7).fill = FILL_PASS
        elif f["status"] == "FAIL":
            ws2.cell(row=row_idx, column=7).fill = FILL_FAIL

    # Auto-filter
    if ws2.max_row > 1:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws2.max_row}"
    for col_idx in range(1, len(columns) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 3: Compliance by Target ──
    ws3 = wb.create_sheet("Compliance by Target")
    ws3.append(["Target", "IP Address", "Benchmark", "Compliance %", "Passed", "Failed", "Errors"])
    for col_idx in range(1, 8):
        ws3.cell(row=1, column=col_idx).font = label_font
    for target in data["targets"]:
        for scan in target["scans"]:
            ws3.append([
                target["hostname"],
                target["ip_address"],
                scan["benchmark_name"],
                scan["compliance_percentage"],
                scan["passed"],
                scan["failed"],
                scan["errors"],
            ])
    for col_idx in range(1, 8):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 4: Compliance by Category ──
    ws4 = wb.create_sheet("Compliance by Category")
    ws4.append(["Category", "Total", "Passed", "Failed", "Compliance %"])
    for col_idx in range(1, 6):
        ws4.cell(row=1, column=col_idx).font = label_font

    # Conditional formatting fills for compliance %
    _FILL_COMP_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _FILL_COMP_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _FILL_COMP_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    by_category = data.get("by_category", {})
    for cat_key in sorted(by_category.keys(), key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else 0, x)):
        info = by_category[cat_key]
        comp = round((info["passed"] / info["total"]) * 100, 1) if info["total"] > 0 else 0.0
        row_idx = ws4.max_row + 1
        ws4.append([info.get("label", f"Section {cat_key}"), info["total"], info["passed"], info["failed"], comp])
        # Conditional fill on compliance cell
        comp_cell = ws4.cell(row=row_idx, column=5)
        if comp >= 80:
            comp_cell.fill = _FILL_COMP_GREEN
        elif comp >= 60:
            comp_cell.fill = _FILL_COMP_YELLOW
        else:
            comp_cell.fill = _FILL_COMP_RED
    for col_idx in range(1, 6):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 30

    # ── Sheet 5: Groups (when builder groups are present) ──
    builder_groups = data.get("builder_groups")
    if builder_groups:
        ws5 = wb.create_sheet("Groups")
        group_cols = ["Group", "Compliance %", "Passed", "Failed", "Total", "Rule Section", "Rule Title", "Description", "Severity", "Status"]
        ws5.append(group_cols)
        for col_idx in range(1, len(group_cols) + 1):
            ws5.cell(row=1, column=col_idx).font = label_font

        # Build rule lookups from findings
        findings_by_rule: dict[int, dict] = {}
        for f in data["findings"]:
            if f.get("_rule_id") and f["_rule_id"] not in findings_by_rule:
                findings_by_rule[f["_rule_id"]] = f

        for bg in builder_groups:
            # Compute group-level compliance
            g_pass = sum(1 for rid in bg["rule_ids"] if findings_by_rule.get(rid, {}).get("status") == "PASS")
            g_fail = sum(1 for rid in bg["rule_ids"] if findings_by_rule.get(rid, {}).get("status") == "FAIL")
            g_total = len(bg["rule_ids"])
            g_comp = round((g_pass / g_total) * 100, 1) if g_total > 0 else 0.0

            for i, rid in enumerate(bg["rule_ids"]):
                fdata = findings_by_rule.get(rid, {})
                row_idx = ws5.max_row + 1
                ws5.append([
                    bg["name"] if i == 0 else "",  # Show group name only on first row
                    g_comp if i == 0 else "",
                    g_pass if i == 0 else "",
                    g_fail if i == 0 else "",
                    g_total if i == 0 else "",
                    fdata.get("section_number", ""),
                    _clean(fdata.get("rule_title", f"Rule #{rid}")),
                    _clean(fdata.get("description", "")[:200]),
                    fdata.get("severity", ""),
                    fdata.get("status", ""),
                ])
                sev_fill = SEVERITY_FILLS.get(fdata.get("severity", ""))
                if sev_fill:
                    ws5.cell(row=row_idx, column=9).fill = sev_fill
                if fdata.get("status") == "PASS":
                    ws5.cell(row=row_idx, column=10).fill = FILL_PASS
                elif fdata.get("status") == "FAIL":
                    ws5.cell(row=row_idx, column=10).fill = FILL_FAIL
                # Compliance conditional fill on first row of each group
                if i == 0:
                    comp_cell = ws5.cell(row=row_idx, column=2)
                    if g_comp >= 80:
                        comp_cell.fill = _FILL_COMP_GREEN
                    elif g_comp >= 60:
                        comp_cell.fill = _FILL_COMP_YELLOW
                    else:
                        comp_cell.fill = _FILL_COMP_RED

        if ws5.max_row > 1:
            ws5.auto_filter.ref = f"A1:{get_column_letter(len(group_cols))}{ws5.max_row}"
        for col_idx in range(1, len(group_cols) + 1):
            ws5.column_dimensions[get_column_letter(col_idx)].width = 22
        ws5.column_dimensions["H"].width = 40  # Description

    # ── Apply text wrapping and column widths to Findings ──
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for col_idx in [8, 9, 10, 11, 12, 13]:  # Description through AI Advice
        for row_idx in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap_align
    ws2.column_dimensions["H"].width = 35  # Description
    ws2.column_dimensions["I"].width = 30  # Actual Output
    ws2.column_dimensions["J"].width = 30  # Expected Output
    ws2.column_dimensions["K"].width = 35  # Remediation
    ws2.column_dimensions["L"].width = 30  # Evaluation
    ws2.column_dimensions["M"].width = 30  # AI Advice
    ws2.column_dimensions["P"].width = 25  # Related Services

    # ── Sheet: Device Profiles (from discovery data) ──
    device_profiles = data.get("device_profiles", {})
    dp_list = list(device_profiles.values())
    if dp_list:
        _FILL_DP_HEADER = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
        ws_dp = wb.create_sheet("Device Profiles")
        dp_cols = ["Hostname", "IP Address", "MAC Address", "OS", "OS Version",
                   "Vendor", "Model", "Domain", "Confidence", "Open Ports",
                   "Detection Method", "Last Seen"]
        ws_dp.append(dp_cols)
        for col_idx in range(1, len(dp_cols) + 1):
            c = ws_dp.cell(row=1, column=col_idx)
            c.font = label_font
            c.fill = _FILL_DP_HEADER

        for dp in dp_list:
            ports_str = ", ".join(
                f"{p.get('service', p.get('port', '?'))}:{p.get('port', '?')}"
                for p in dp.get("open_ports", [])
            )
            ws_dp.append([
                dp.get("hostname", ""),
                dp.get("ip_address", ""),
                dp.get("mac_address", ""),
                dp.get("os_guess", ""),
                dp.get("os_version", ""),
                dp.get("vendor", ""),
                dp.get("device_model", ""),
                dp.get("domain", ""),
                dp.get("confidence", 0),
                ports_str,
                dp.get("detection_method", ""),
                dp.get("last_seen", "")[:10] if dp.get("last_seen") else "",
            ])

        for col_idx in range(1, len(dp_cols) + 1):
            ws_dp.column_dimensions[get_column_letter(col_idx)].width = 20
        ws_dp.column_dimensions["J"].width = 40  # Open Ports

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# CSV generation
# ---------------------------------------------------------------------------

def generate_csv_report(data: dict, include_passed: bool = True) -> str:
    """Flat CSV export of all findings with auditor fields."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "scan_id", "target", "benchmark", "rule_section", "rule_title",
        "description", "severity", "status",
        "actual_output", "expected_output", "default_value", "rationale",
        "remediation", "evaluation_explanation", "ai_advice",
        "auditor_notes", "auditor_override",
    ])
    for f in data["findings"]:
        if not include_passed and f["status"] == "PASS":
            continue
        writer.writerow([
            f["scan_id"],
            f["target_hostname"],
            f["benchmark_name"],
            f["section_number"],
            f["rule_title"],
            f.get("description", ""),
            f["severity"],
            f["status"],
            f["actual_output"],
            f["expected_output"],
            f.get("default_value", ""),
            f.get("rationale", ""),
            f["remediation"],
            f.get("evaluation_explanation", ""),
            f.get("ai_advice", ""),
            f.get("auditor_notes", ""),
            f.get("auditor_override", ""),
        ])
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

def generate_html_report(data: dict, include_passed: bool, db: Session | None = None) -> str:
    """Self-contained interactive HTML dashboard — pure inline SVG charts, zero external deps."""
    findings = data["findings"]
    if not include_passed:
        findings = [f for f in findings if f["status"] != "PASS"]

    # ── White-label settings ──
    _wl = _load_whitelabel_settings(db) if db else {}
    company_name = _wl.get("company_name", "AuditForge")
    company_logo_base64 = _wl.get("company_logo_base64", "")
    auditor_name = _wl.get("auditor_name", "")

    # ── Phase 2: grouped findings & builder metadata ──
    grouped_findings = _build_grouped_findings(data, findings)
    sections = data.get("sections") or {}
    group_summaries = data.get("group_summaries") or {}
    audience = data.get("audience", "technical")
    has_builder = bool(data.get("builder_groups"))

    # ── Generate all SVG charts (shared logic) ──
    charts = _generate_all_charts(data, grouped_findings)

    # ── Compute risk score (parity with PDF) ──
    risk_score = _compute_risk_score(data)

    # ── Per-target findings breakdown (parity with PDF) ──
    per_target_findings = _build_per_target_findings(data)

    # ── Device profiles list (parity with PDF) ──
    dp = data.get("device_profiles", {})
    device_profiles_list = list(dp.values())
    has_device_profiles = len(device_profiles_list) > 0

    # ── Convert AI summary from Markdown to HTML ──
    ai_summary_html = _md_to_html(data.get("ai_summary", ""))

    # ── Phase 4: Business Impact Section ──
    business_impact_groups = _generate_business_impact(findings)

    # ── Findings JSON for JS filtering/sorting engine ──
    findings_json = json.dumps(findings, default=str).replace("</", "<\\/")

    fp_summary = data.get("fp_summary", {})

    env = Environment(loader=FileSystemLoader(TEMPLATES_DIR), autoescape=True)
    template = env.get_template("report_dashboard.html.j2")

    return template.render(
        title=data.get("title") or f"Audit Report - {data.get('mission_name', '')}",
        client_name=data.get("client_name", ""),
        mission_name=data.get("mission_name", ""),
        benchmark_name=data.get("benchmark_name", "CIS Benchmark"),
        company_name=company_name,
        company_logo_base64=company_logo_base64,
        auditor_name=auditor_name,
        date_range=data.get("date_range", ""),
        generated_at=data.get("generated_at", ""),
        ai_summary=data.get("ai_summary", ""),
        ai_summary_html=ai_summary_html,
        summary=data["summary"],
        targets=data.get("targets", []),
        findings=findings,
        by_category=data.get("by_category", {}),
        by_target=data.get("by_target", []),
        categories_enriched=data.get("categories_enriched", []),
        risk_score=risk_score,
        per_target_findings=per_target_findings,
        fp_summary=fp_summary,
        findings_json=findings_json,
        grouped_findings=grouped_findings,
        section_toggles=sections,
        group_summaries=group_summaries,
        audience=audience,
        has_builder=has_builder,
        is_finalized=data.get("is_finalized", False),
        audit_info=data.get("audit_info", {}),
        device_profiles=data.get("device_profiles", {}),
        device_profiles_list=device_profiles_list,
        has_device_profiles=has_device_profiles,
        include_passed=include_passed,
        business_impact_groups=business_impact_groups,
        **charts,
    )



# ---------------------------------------------------------------------------
# AI executive summary
# ---------------------------------------------------------------------------

async def generate_ai_summary(data: dict) -> str:
    """Use the LLM to produce a structured executive summary of the audit data."""
    summary = data["summary"]
    compliance = summary["overall_compliance"]

    # ---- severity breakdown ----
    sev_lines = "\n".join(
        f"  - {sev}: {info['total']} total, {info['passed']} passed, {info['failed']} failed "
        f"({round(info['passed'] / info['total'] * 100) if info['total'] else 0}% compliant)"
        for sev, info in summary["by_severity"].items()
    )

    # ---- category breakdown (top 5 worst) ----
    cats = data.get("by_category", {})
    worst_cats = sorted(
        cats.items(),
        key=lambda x: x[1].get("failed", 0),
        reverse=True,
    )[:5]
    cat_lines = "\n".join(
        f"  - {cat}: {info.get('failed', 0)} failures out of {info.get('total', 0)} rules"
        for cat, info in worst_cats
    ) if worst_cats else "  (no category data available)"

    # ---- top failures with more context ----
    top_failures = [f for f in data["findings"] if f["status"] == "FAIL"][:15]
    failure_lines = "\n".join(
        f"  - [{f['severity'].upper()}] {f['section_number']} {f['rule_title']}"
        + (f"  — {f['description'][:120]}" if f.get("description") else "")
        for f in top_failures
    )

    # ---- risk level label ----
    if compliance >= 90:
        risk_label = "LOW"
    elif compliance >= 70:
        risk_label = "MODERATE"
    elif compliance >= 50:
        risk_label = "HIGH"
    else:
        risk_label = "CRITICAL"

    # ---- false-positive context ----
    fp_count = sum(
        1 for f in data["findings"]
        if f.get("auditor_status") in ("false_positive", "False Positive")
    )
    fp_note = (
        f"\n\nFalse Positives Flagged by Auditor: {fp_count}"
        if fp_count else ""
    )

    prompt = f"""You are a senior cybersecurity auditor writing an executive summary for a CIS benchmark configuration audit report.

──────────────────────────────────────────
AUDIT CONTEXT
──────────────────────────────────────────
Client:       {data.get('client_name', 'N/A')}
Mission:      {data.get('mission_name', 'N/A')}
Date Range:   {data.get('date_range', 'N/A')}
Benchmark:    {data.get('benchmark_name', 'CIS Benchmark')}
Target(s):    {', '.join(t.get('hostname', 'Unknown') for t in data.get('targets', [])) or 'N/A'}

──────────────────────────────────────────
KEY METRICS
──────────────────────────────────────────
Overall Compliance: {compliance}%  (Risk Level: {risk_label})
Total Rules:        {summary['total_rules']}
Passed:             {summary['passed']}
Failed:             {summary['failed']}
Errors:             {summary['errors']}{fp_note}

By Severity:
{sev_lines}

──────────────────────────────────────────
WORST CATEGORIES (by failure count)
──────────────────────────────────────────
{cat_lines}

──────────────────────────────────────────
TOP {len(top_failures)} FAILED RULES
──────────────────────────────────────────
{failure_lines}

──────────────────────────────────────────
INSTRUCTIONS
──────────────────────────────────────────
Write the executive summary using this EXACT structure (use these headings):

1. **Overview** — One paragraph stating the audit scope, benchmark, client, date,
   and the overall compliance percentage with the risk label.

2. **Key Findings** — Two paragraphs highlighting the most significant failures
   grouped by theme (e.g., authentication weaknesses, logging gaps, network
   exposure). Reference specific rule numbers.

3. **Risk Assessment** — One paragraph assessing the overall security posture,
   relating the compliance score to real-world risk. Mention which severity
   categories are most concerning.

4. **Recommendations** — A numbered list of 5-8 prioritised remediation actions,
   starting with quick wins and ending with strategic improvements. Each item
   should reference the related category or rule(s).

5. **Conclusion** — One short paragraph with a professional closing statement
   about next steps and timeline expectations.

TONE: Professional, objective, factual. Avoid marketing language.
LENGTH: 500-800 words total.
FORMAT: Use Markdown headings (##) and bullet/numbered lists."""

    system_prompt = (
        "You are a cybersecurity audit report writer specialising in CIS benchmark assessments. "
        "Produce clear, structured, professional executive summaries suitable for C-level "
        "and IT management audiences. Use precise language, cite specific rule numbers, "
        "and provide actionable recommendations."
    )

    try:
        return await llm_manager.invoke(prompt, system_prompt=system_prompt, task="reports")
    except Exception:
        logger.exception("AI summary generation failed")
        return "AI executive summary could not be generated. Please check LLM configuration."
